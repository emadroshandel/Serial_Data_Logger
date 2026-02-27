"""
Serial Data Logger
==================
A Python GUI application for serial port communication.

Continuous Save modes
─────────────────────
① Every reading    — records all data as it arrives
② Simple window    — records for <window>s every <period>s
③ Pre/Post window  — at each trigger (every <period>s), flush the last
                     <pre>s from the rolling buffer THEN record live for
                     <post>s.  Gives you context before AND after each event.

Requirements:  pip install pyserial openpyxl
Run:           python serial_logger.py

Author: Emad Roshandel
"""

import collections
import csv
import datetime
import os
import re
import threading
import time
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

import openpyxl
import serial
import serial.tools.list_ports
from openpyxl.styles import Alignment, Font, PatternFill

# ─────────────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────────────

_ILLEGAL = re.compile(
    r"[\x00-\x08\x0b-\x0c\x0e-\x1f\x7f-\x84\x86-\x9f\ud800-\udfff\ufffe\uffff]"
)

def sanitize(v: str) -> str:
    return _ILLEGAL.sub("", v)


def csv_to_excel(csv_path: str, xlsx_path: str) -> int:
    """
    CSV columns: index | full_timestamp | date | time | value | phase_flag
    phase_flag: 'pre' | 'post' | '1' (simple window) | '' (every reading)
    """
    FILLS = {
        "pre":  PatternFill("solid", fgColor="C8E6C9"),   # soft green  = pre-buffer
        "post": PatternFill("solid", fgColor="FFE0B2"),   # soft orange = post/live
        "1":    PatternFill("solid", fgColor="FFE0B2"),   # simple window same as post
        "alt":  PatternFill("solid", fgColor="D6E4F0"),   # alternating blue
        "":     PatternFill(),
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Serial Data"

    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True, size=11)
    for col, (h, w) in enumerate(
        zip(["#", "Timestamp", "Date", "Time", "Raw Value"], [6, 26, 14, 14, 30]),
        start=1
    ):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 20

    count = 0
    with open(csv_path, newline="", encoding="utf-8") as f:
        for i, row in enumerate(csv.reader(f), start=1):
            if len(row) < 5:
                continue
            flag = row[5].strip() if len(row) >= 6 else ""
            if flag in ("pre", "post", "1"):
                fill = FILLS[flag]
            else:
                fill = FILLS["alt"] if i % 2 == 0 else FILLS[""]
            for col, v in enumerate(row[:5], start=1):
                val = int(v) if col == 1 and v.strip().isdigit() else v
                c = ws.cell(row=i + 1, column=col, value=val)
                c.fill = fill
                c.alignment = Alignment(horizontal="center" if col == 1 else "left")
            count += 1

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Continuous Save — Export Summary"
    ws2["A1"].font = Font(bold=True, size=13)
    for r, (k, v) in enumerate([
        ("Export Time",   datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Records", count),
        ("Green rows",    "Pre-buffer (data before trigger)"),
        ("Orange rows",   "Post-window (data after trigger)"),
    ], start=2):
        ws2.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=v)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 45
    wb.save(xlsx_path)
    return count


# ─────────────────────────────────────────────────────────────────────────────
#  Serial Logger Core
# ─────────────────────────────────────────────────────────────────────────────

class SerialLogger:
    """
    Three continuous-save modes
    ───────────────────────────
    A  every reading   window=0, period=0, pre=0, post=0
    B  simple window   window>0, period>0, pre=0, post=0
       records `window` s live every `period` s
    C  pre/post        pre>0 or post>0, period>0
       at each trigger: flush last `pre` s from rolling buffer,
       then record live for `post` s
    """

    def __init__(self):
        self.serial_port = None
        self.is_running  = False
        # Rolling buffer — kept trimmed to max(pre) seconds depth
        self.data_buffer = []          # [(datetime, str)]  all readings ever
        self.lock        = threading.Lock()

        # CS shared
        self._cs_active   = False
        self._cs_file     = None
        self._cs_writer   = None
        self._cs_csv_path = ""
        self._cs_index    = 1
        self._cs_lock     = threading.Lock()
        self._cs_stop_evt = threading.Event()

        # Mode B  (simple window)
        self._cs_gate     = True
        self._cs_window   = 0.0
        self._cs_period   = 0.0

        # Mode C  (pre/post)
        self._cs_pre      = 0.0
        self._cs_post     = 0.0
        self._cs_in_post  = False   # True while recording the post-window live

        self._cs_on_gate_cb = None  # callback(phase: str)  phase = 'pre'|'post'|'wait'|'open'|'closed'

        # Nth-sample decimation (applies on top of all CS modes)
        # If _nth_step > 1, only every Nth incoming reading is written to CSV.
        self._nth_step    = 1   # 1 = save every reading (disabled)
        self._nth_counter = 0   # counts incoming readings since last save

    # ── Connection ───────────────────────────────────────────────────────────

    def connect(self, port, baud):
        self.serial_port = serial.Serial(port, baud, timeout=1)

    def disconnect(self):
        self.is_running = False
        self.stop_continuous_save()
        if self.serial_port and self.serial_port.is_open:
            self.serial_port.close()

    # ── Reading loop ─────────────────────────────────────────────────────────

    def start_reading(self, interval_s, on_data_cb):
        self.is_running = True
        threading.Thread(
            target=self._read_loop, args=(interval_s, on_data_cb), daemon=True
        ).start()

    def _read_loop(self, interval_s, on_data_cb):
        nxt = time.time()
        while self.is_running:
            now = time.time()
            if now >= nxt:
                nxt = now + interval_s
                try:
                    if self.serial_port and self.serial_port.in_waiting > 0:
                        raw = (self.serial_port.readline()
                               .decode("utf-8", errors="replace").strip())
                        if raw:
                            ts    = datetime.datetime.now()
                            clean = sanitize(raw)
                            with self.lock:
                                self.data_buffer.append((ts, clean))
                            saved = self._cs_write(ts, clean)
                            on_data_cb(ts, clean, saved)
                except Exception as e:
                    on_data_cb(None, f"[ERROR] {e}", False)
                    break
            time.sleep(0.005)

    # ── Snapshot save ────────────────────────────────────────────────────────

    def save_snapshot_to_excel(self, filepath, interval_s):
        with self.lock:
            cutoff = datetime.datetime.now() - datetime.timedelta(seconds=interval_s)
            rows   = [(ts, v) for ts, v in self.data_buffer if ts >= cutoff]
        if not rows:
            return 0

        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Serial Data"
        hf = PatternFill("solid", fgColor="1F4E79")
        hfont = Font(color="FFFFFF", bold=True, size=11)
        for col, (h, w) in enumerate(
            zip(["#","Timestamp","Date","Time","Raw Value"], [6,26,14,14,30]), start=1
        ):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hf; c.font = hfont
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[c.column_letter].width = w
        ws.row_dimensions[1].height = 20

        af = PatternFill("solid", fgColor="D6E4F0")
        for i, (ts, val) in enumerate(rows, start=1):
            fill = af if i % 2 == 0 else PatternFill()
            for col, v in enumerate([
                i, ts.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
                ts.strftime("%Y-%m-%d"), ts.strftime("%H:%M:%S.%f")[:-3], val
            ], start=1):
                c = ws.cell(row=i+1, column=col, value=v)
                c.fill = fill
                c.alignment = Alignment(horizontal="center" if col==1 else "left")

        ws2 = wb.create_sheet("Summary"); ws2["A1"] = "Snapshot"
        ws2["A1"].font = Font(bold=True, size=13)
        for r, (k, v) in enumerate([
            ("Export Time",   datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Interval (s)",  interval_s),
            ("Total Records", len(rows)),
            ("First",         rows[0][0].strftime("%H:%M:%S.%f")[:-3]),
            ("Last",          rows[-1][0].strftime("%H:%M:%S.%f")[:-3]),
        ], start=2):
            ws2.cell(row=r, column=1, value=k).font = Font(bold=True)
            ws2.cell(row=r, column=2, value=v)
        ws2.column_dimensions["A"].width = 20; ws2.column_dimensions["B"].width = 30
        wb.save(filepath)
        return len(rows)

    # ── Continuous save: public API ───────────────────────────────────────────

    def start_continuous_save(self, csv_path, window=0.0, period=0.0,
                               pre=0.0, post=0.0, on_gate_cb=None):
        with self._cs_lock:
            if self._cs_active:
                return
            f = open(csv_path, "w", newline="", encoding="utf-8")
            self._cs_file       = f
            self._cs_writer     = csv.writer(f)
            self._cs_csv_path   = csv_path
            self._cs_index      = 1
            self._cs_window     = window
            self._cs_period     = period
            self._cs_pre        = pre
            self._cs_post       = post
            self._cs_in_post    = False
            self._cs_gate       = True
            self._cs_on_gate_cb = on_gate_cb
            self._cs_active     = True
            self._cs_stop_evt.clear()

        # Pick scheduler thread based on mode
        if pre > 0 or post > 0:
            # Mode C — pre/post
            t = threading.Thread(target=self._scheduler_prepost, daemon=True)
        elif window > 0 and period > 0:
            # Mode B — simple window
            t = threading.Thread(target=self._scheduler_window, daemon=True)
        else:
            # Mode A — every reading, no scheduler needed
            return
        t.start()

    def stop_continuous_save(self) -> str:
        with self._cs_lock:
            if not self._cs_active:
                return ""
            self._cs_active = False
            self._cs_stop_evt.set()
            path = self._cs_csv_path
            try:
                self._cs_file.flush()
                self._cs_file.close()
            except Exception:
                pass
            self._cs_file = self._cs_writer = None
            return path

    # ── Continuous save: write one row ───────────────────────────────────────

    def _cs_write(self, ts, val, flag="") -> bool:
        """Write row if gate is open (mode A/B) or in post phase (mode C)."""
        with self._cs_lock:
            if not self._cs_active or self._cs_writer is None:
                return False
            # ── Step 1: gate / mode check ────────────────────────────────────
            # Mode C: only write when in post-window
            if self._cs_pre > 0 or self._cs_post > 0:
                if not self._cs_in_post:
                    return False
                flag = "post"
            # Mode B: check gate
            elif not self._cs_gate:
                return False
            elif self._cs_window > 0:
                flag = "1"
            # ── Step 2: Nth-sample decimation ────────────────────────────────
            # Count only readings that pass the gate, so N=10 always means
            # "1 saved per 10 gate-passing readings", regardless of mode.
            if self._nth_step > 1:
                self._nth_counter += 1
                if self._nth_counter < self._nth_step:
                    return False
                self._nth_counter = 0   # reset after saving
            try:
                self._cs_writer.writerow([
                    self._cs_index,
                    ts.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
                    ts.strftime("%Y-%m-%d"),
                    ts.strftime("%H:%M:%S.%f")[:-3],
                    val, flag,
                ])
                self._cs_file.flush()
                self._cs_index += 1
                return True
            except Exception:
                return False

    def _cs_write_row_direct(self, ts, val, flag):
        """Write a row unconditionally (used to flush pre-buffer)."""
        with self._cs_lock:
            if not self._cs_active or self._cs_writer is None:
                return
            try:
                self._cs_writer.writerow([
                    self._cs_index,
                    ts.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
                    ts.strftime("%Y-%m-%d"),
                    ts.strftime("%H:%M:%S.%f")[:-3],
                    val, flag,
                ])
                self._cs_file.flush()
                self._cs_index += 1
            except Exception:
                pass

    # ── Scheduler: Mode B (simple window) ────────────────────────────────────

    def _scheduler_window(self):
        stop = self._cs_stop_evt
        while True:
            # OPEN
            self._cs_gate = True
            if self._cs_on_gate_cb:
                self._cs_on_gate_cb("open")
            if stop.wait(timeout=self._cs_window):
                break
            # CLOSED
            self._cs_gate = False
            if self._cs_on_gate_cb:
                self._cs_on_gate_cb("closed")
            gap = self._cs_period - self._cs_window
            if gap > 0:
                if stop.wait(timeout=gap):
                    break
            if not self._cs_active:
                break

    # ── Scheduler: Mode C (pre/post) ─────────────────────────────────────────

    def _scheduler_prepost(self):
        """
        Every `period` seconds:
          1. Grab the last `pre` seconds from data_buffer → write with flag='pre'
          2. Open the gate for `post` seconds            → _cs_write flags as 'post'
          3. Close gate, sleep for remainder, repeat
        """
        stop = self._cs_stop_evt
        pre  = self._cs_pre
        post = self._cs_post

        # Wait for the first trigger (one full period)
        if self._cs_on_gate_cb:
            self._cs_on_gate_cb("wait")
        if stop.wait(timeout=self._cs_period):
            return

        while self._cs_active:
            # ── 1. Flush pre-buffer ───────────────────────────────────────────
            if pre > 0:
                cutoff = datetime.datetime.now() - datetime.timedelta(seconds=pre)
                with self.lock:
                    pre_rows = [(ts, v) for ts, v in self.data_buffer if ts >= cutoff]
                if self._cs_on_gate_cb:
                    self._cs_on_gate_cb("pre")
                for ts, v in pre_rows:
                    self._cs_write_row_direct(ts, v, "pre")

            # ── 2. Open post-window ───────────────────────────────────────────
            self._cs_in_post = True
            if self._cs_on_gate_cb:
                self._cs_on_gate_cb("post")
            if post > 0:
                if stop.wait(timeout=post):
                    break

            # ── 3. Close and sleep ────────────────────────────────────────────
            self._cs_in_post = False
            if self._cs_on_gate_cb:
                self._cs_on_gate_cb("wait")
            gap = self._cs_period - post   # remaining time until next trigger
            if gap > 0:
                if stop.wait(timeout=gap):
                    break

        self._cs_in_post = False


# ─────────────────────────────────────────────────────────────────────────────
#  GUI
# ─────────────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Serial Data Logger")
        self.geometry("980x730")
        self.resizable(True, True)
        self.configure(bg="#1a1a2e")

        self.logger        = SerialLogger()
        self.auto_save_job = None
        self.output_dir    = os.path.expanduser("~")
        self._cs_running   = False
        self._live_index   = 0
        self._saved_count  = 0

        self._build_ui()
        self._refresh_ports()

    # ─────────────────────────────────────────────────────────────────────────
    #  UI
    # ─────────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        DARK = "#1a1a2e"; PANEL = "#16213e"; ACCENT = "#0f3460"
        GREEN = "#4ecca3"; RED = "#e94560"; ORANGE = "#f0a500"; TEXT = "#eaeaea"
        MONO = ("Consolas", 10)

        s = ttk.Style(self); s.theme_use("clam")
        s.configure("TLabel",            background=PANEL,    foreground=TEXT,    font=("Segoe UI", 10))
        s.configure("TFrame",            background=PANEL)
        s.configure("TLabelframe",       background=PANEL,    foreground=GREEN,   font=("Segoe UI", 10, "bold"))
        s.configure("TLabelframe.Label", background=PANEL,    foreground=GREEN)
        s.configure("TCombobox",         fieldbackground=ACCENT, background=ACCENT, foreground=TEXT)
        s.configure("TEntry",            fieldbackground=ACCENT, foreground=TEXT,  insertcolor=TEXT)
        s.configure("Connect.TButton",   background=GREEN,    foreground=DARK,    font=("Segoe UI", 10, "bold"), padding=6)
        s.configure("Danger.TButton",    background=RED,      foreground="white", font=("Segoe UI", 10, "bold"), padding=6)
        s.configure("Action.TButton",    background=ACCENT,   foreground=GREEN,   font=("Segoe UI", 10), padding=6)
        s.configure("CSStart.TButton",   background="#1a6b45", foreground="white", font=("Segoe UI", 10, "bold"), padding=6)
        s.configure("CSStop.TButton",    background=ORANGE,   foreground=DARK,    font=("Segoe UI", 10, "bold"), padding=6)
        s.map("Connect.TButton", background=[("active","#3ab087")])
        s.map("Danger.TButton",  background=[("active","#c73652")])
        s.map("Action.TButton",  background=[("active","#0d2a4a")])
        s.map("CSStart.TButton", background=[("active","#155c39")])
        s.map("CSStop.TButton",  background=[("active","#c98600")])

        # Top bar
        top = tk.Frame(self, bg="#0f3460", pady=10); top.pack(fill="x")
        tk.Label(top, text="⚡ Serial Data Logger", bg="#0f3460", fg=GREEN,
                 font=("Segoe UI", 16, "bold")).pack(side="left", padx=16)
        self.status_label = tk.Label(top, text="● Disconnected", bg="#0f3460", fg=RED,
                                     font=("Segoe UI", 11, "bold"))
        self.status_label.pack(side="right", padx=16)
        tk.Label(top, text="made by Emad", bg="#0f3460", fg="#4a7fa5",
                 font=("Segoe UI", 9, "italic")).pack(side="right", padx=(0, 8))

        main = tk.Frame(self, bg=DARK); main.pack(fill="both", expand=True, padx=10, pady=8)
        left = tk.Frame(main, bg=DARK); left.pack(side="left", fill="y", padx=(0,8))

        # ── Connection ────────────────────────────────────────────────────────
        conn = ttk.LabelFrame(left, text=" Connection ", padding=10)
        conn.pack(fill="x", pady=(0,8))
        ttk.Label(conn, text="Port").grid(row=0, column=0, sticky="w", pady=3)
        pr = tk.Frame(conn, bg=PANEL); pr.grid(row=0, column=1, sticky="ew", pady=3, columnspan=2)
        self.port_var = tk.StringVar()
        self.port_cb  = ttk.Combobox(pr, textvariable=self.port_var, width=13, state="readonly")
        self.port_cb.pack(side="left")
        ttk.Button(pr, text="↻", width=3, style="Action.TButton",
                   command=self._refresh_ports).pack(side="left", padx=4)
        ttk.Label(conn, text="Baud Rate").grid(row=1, column=0, sticky="w", pady=3)
        self.baud_var = tk.StringVar(value="9600")
        ttk.Combobox(conn, textvariable=self.baud_var, width=16, state="readonly",
                     values=["300","1200","2400","4800","9600","19200",
                             "38400","57600","115200","230400","460800","921600"]
                     ).grid(row=1, column=1, sticky="ew", pady=3, columnspan=2)
        conn.columnconfigure(1, weight=1)
        self.connect_btn = ttk.Button(conn, text="Connect", style="Connect.TButton",
                                      command=self._toggle_connection)
        self.connect_btn.grid(row=2, column=0, columnspan=3, pady=(8,0))

        # ── Sampling ──────────────────────────────────────────────────────────
        samp = ttk.LabelFrame(left, text=" Sampling ", padding=10); samp.pack(fill="x", pady=(0,8))
        ttk.Label(samp, text="Sample Rate (Hz)").grid(row=0, column=0, sticky="w", pady=3)
        self.rate_var = tk.StringVar(value="10")
        ttk.Entry(samp, textvariable=self.rate_var, width=10).grid(row=0, column=1, sticky="ew", pady=3)
        ttk.Label(samp, text="Snapshot Interval (s)").grid(row=1, column=0, sticky="w", pady=3)
        self.interval_var = tk.StringVar(value="60")
        ttk.Entry(samp, textvariable=self.interval_var, width=10).grid(row=1, column=1, sticky="ew", pady=3)
        ttk.Label(samp, text="Auto-Save Every (s)").grid(row=2, column=0, sticky="w", pady=3)
        self.autosave_var = tk.StringVar(value="0")
        ttk.Entry(samp, textvariable=self.autosave_var, width=10).grid(row=2, column=1, sticky="ew", pady=3)
        ttk.Label(samp, text="(0=off)").grid(row=2, column=2, sticky="w", padx=4)
        ttk.Label(samp, text="Save every N inputs").grid(row=3, column=0, sticky="w", pady=3)
        self.nth_var = tk.StringVar(value="1")
        ttk.Entry(samp, textvariable=self.nth_var, width=10).grid(row=3, column=1, sticky="ew", pady=3)
        ttk.Label(samp, text="(1=all, 10=every 10th)").grid(row=3, column=2, sticky="w", padx=4)
        samp.columnconfigure(1, weight=1)

        # ── Output ────────────────────────────────────────────────────────────
        out = ttk.LabelFrame(left, text=" Output ", padding=10); out.pack(fill="x", pady=(0,8))
        ttk.Label(out, text="Save Folder").grid(row=0, column=0, sticky="w", pady=3)
        self.dir_var = tk.StringVar(value=self.output_dir)
        ttk.Entry(out, textvariable=self.dir_var, width=20).grid(row=0, column=1, sticky="ew", pady=3)
        ttk.Button(out, text="…", width=3, style="Action.TButton",
                   command=self._browse_dir).grid(row=0, column=2, padx=4)
        ttk.Label(out, text="Filename Prefix").grid(row=1, column=0, sticky="w", pady=3)
        self.prefix_var = tk.StringVar(value="serial_log")
        ttk.Entry(out, textvariable=self.prefix_var, width=20).grid(row=1, column=1, sticky="ew", pady=3)
        out.columnconfigure(1, weight=1)

        # ── Continuous Save ───────────────────────────────────────────────────
        cs = ttk.LabelFrame(left, text=" Continuous Save ", padding=10); cs.pack(fill="x", pady=(0,8))

        self.cs_mode_var = tk.StringVar(value="all")

        def rb(text, value, row):
            tk.Radiobutton(cs, text=text, variable=self.cs_mode_var, value=value,
                           bg=PANEL, fg=TEXT, selectcolor=ACCENT,
                           activebackground=PANEL, activeforeground=GREEN,
                           font=("Segoe UI", 10),
                           command=self._on_cs_mode_change
                           ).grid(row=row, column=0, columnspan=3, sticky="w", pady=2)

        rb("① Continuous (no time window)",   "all",     0)
        rb("② Simple window",          "window",  1)
        rb("③ Pre / Post window",       "prepost", 2)

        # ── Mode B settings ───────────────────────────────────────────────────
        self.win_frame = tk.Frame(cs, bg=PANEL)
        self.win_frame.grid(row=3, column=0, columnspan=3, sticky="ew", padx=8, pady=(2,6))

        ttk.Label(self.win_frame, text="Record for").grid(row=0, column=0, sticky="w", pady=2)
        self.cs_window_var = tk.StringVar(value="5")
        ttk.Entry(self.win_frame, textvariable=self.cs_window_var, width=7).grid(row=0, column=1, padx=4)
        ttk.Label(self.win_frame, text="s  every").grid(row=0, column=2, sticky="w")
        self.cs_period_var = tk.StringVar(value="30")
        ttk.Entry(self.win_frame, textvariable=self.cs_period_var, width=7).grid(row=0, column=3, padx=4)
        ttk.Label(self.win_frame, text="s").grid(row=0, column=4, sticky="w")
        self.win_frame.grid_remove()

        # ── Mode C settings ───────────────────────────────────────────────────
        self.pp_frame = tk.Frame(cs, bg=PANEL)
        self.pp_frame.grid(row=3, column=0, columnspan=3, sticky="ew", padx=8, pady=(2,6))

        ttk.Label(self.pp_frame, text="Trigger every").grid(row=0, column=0, sticky="w", pady=2)
        self.cs_pp_period_var = tk.StringVar(value="120")
        ttk.Entry(self.pp_frame, textvariable=self.cs_pp_period_var, width=7).grid(row=0, column=1, padx=4)
        ttk.Label(self.pp_frame, text="s").grid(row=0, column=2, sticky="w")

        ttk.Label(self.pp_frame, text="Pre-buffer").grid(row=1, column=0, sticky="w", pady=2)
        self.cs_pre_var = tk.StringVar(value="20")
        ttk.Entry(self.pp_frame, textvariable=self.cs_pre_var, width=7).grid(row=1, column=1, padx=4)
        ttk.Label(self.pp_frame, text="s before trigger").grid(row=1, column=2, sticky="w")

        ttk.Label(self.pp_frame, text="Post-window").grid(row=2, column=0, sticky="w", pady=2)
        self.cs_post_var = tk.StringVar(value="30")
        ttk.Entry(self.pp_frame, textvariable=self.cs_post_var, width=7).grid(row=2, column=1, padx=4)
        ttk.Label(self.pp_frame, text="s after trigger").grid(row=2, column=2, sticky="w")

        # Colour legend
        leg = tk.Frame(self.pp_frame, bg=PANEL); leg.grid(row=3, column=0, columnspan=3, sticky="w", pady=(6,0))
        tk.Label(leg, text="■", bg=PANEL, fg="#4caf50", font=("Segoe UI",12)).pack(side="left")
        tk.Label(leg, text=" pre  ", bg=PANEL, fg="#aaaaaa", font=("Segoe UI",9)).pack(side="left")
        tk.Label(leg, text="■", bg=PANEL, fg="#f0a500", font=("Segoe UI",12)).pack(side="left")
        tk.Label(leg, text=" post", bg=PANEL, fg="#aaaaaa", font=("Segoe UI",9)).pack(side="left")

        self.pp_frame.grid_remove()

        # Start / Stop
        self.cs_start_btn = ttk.Button(cs, text="▶  Start Continuous Save",
                                       style="CSStart.TButton",
                                       command=self._start_continuous_save)
        self.cs_start_btn.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(4,2))

        self.cs_stop_btn = ttk.Button(cs, text="■  Stop & Save Excel",
                                      style="CSStop.TButton",
                                      command=self._stop_continuous_save)
        self.cs_stop_btn.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(0,4))
        self.cs_stop_btn.grid_remove()

        self.cs_status = tk.Label(cs, text="", bg=PANEL, fg=ORANGE,
                                  font=("Segoe UI",9,"italic"), wraplength=230, justify="left")
        self.cs_status.grid(row=6, column=0, columnspan=3, sticky="w")

        self.gate_label = tk.Label(cs, text="", bg=PANEL, font=("Segoe UI",10,"bold"), width=24)
        self.gate_label.grid(row=7, column=0, columnspan=3, sticky="w", pady=(2,0))
        cs.columnconfigure(1, weight=1)

        # ── Other actions ─────────────────────────────────────────────────────
        act = tk.Frame(left, bg=DARK); act.pack(fill="x", pady=4)
        ttk.Button(act, text="💾  Save Snapshot to Excel",
                   style="Action.TButton", command=self._save_snapshot).pack(fill="x", pady=(0,4))
        ttk.Separator(act).pack(fill="x", pady=4)
        ttk.Button(act, text="🗑  Clear Live View",
                   style="Action.TButton", command=self._clear_log).pack(fill="x", pady=2)
        ttk.Button(act, text="⏹  Stop / Disconnect",
                   style="Danger.TButton", command=self._disconnect).pack(fill="x", pady=2)

        self.stats_label = tk.Label(left, text="Records: 0  |  Saved: 0",
                                    bg=DARK, fg=GREEN, font=("Segoe UI",10))
        self.stats_label.pack(anchor="w", pady=(6,0))

        # ── Live view ─────────────────────────────────────────────────────────
        right = tk.Frame(main, bg=DARK); right.pack(side="left", fill="both", expand=True)

        hdr = tk.Frame(right, bg="#0f3460"); hdr.pack(fill="x", pady=(0,2))
        tk.Label(hdr, text="  #",    bg="#0f3460", fg=GREEN, font=("Consolas",10,"bold"), width=6,  anchor="w").pack(side="left")
        tk.Label(hdr, text="Time",   bg="#0f3460", fg=GREEN, font=("Consolas",10,"bold"), width=16, anchor="w").pack(side="left")
        tk.Label(hdr, text="Value",  bg="#0f3460", fg=GREEN, font=("Consolas",10,"bold"), anchor="w").pack(side="left", fill="x", expand=True)

        self.log_box = scrolledtext.ScrolledText(
            right, bg="#0d1117", fg="#4ecca3", font=MONO,
            insertbackground="white", relief="flat", bd=0,
            state="disabled", wrap="none"
        )
        self.log_box.pack(fill="both", expand=True)
        self.log_box.tag_config("idx",   foreground="#888888")
        self.log_box.tag_config("ts",    foreground="#5b8dd9")
        self.log_box.tag_config("val",   foreground="#4ecca3")
        self.log_box.tag_config("pre",   foreground="#4caf50",  font=(*MONO, "bold"))
        self.log_box.tag_config("post",  foreground="#f0a500",  font=(*MONO, "bold"))
        self.log_box.tag_config("win",   foreground="#f0a500",  font=(*MONO, "bold"))
        self.log_box.tag_config("error", foreground="#e94560")
        self.log_box.tag_config("info",  foreground="#f0a500")

    # ── CS mode radio ─────────────────────────────────────────────────────────

        # ── Footer ───────────────────────────────────────────────────────────
        footer = tk.Frame(self, bg="#0a0a1a", pady=4)
        footer.pack(fill="x", side="bottom")
        tk.Label(footer, text="Made by Emad", bg="#0a0a1a", fg="#4a7fa5",
                 font=("Segoe UI", 9, "italic")).pack(side="right", padx=16)
        tk.Label(footer, text="Serial Data Logger  v1.0", bg="#0a0a1a", fg="#333355",
                 font=("Segoe UI", 9)).pack(side="left", padx=16)

    def _on_cs_mode_change(self):
        m = self.cs_mode_var.get()
        self.win_frame.grid_remove()
        self.pp_frame.grid_remove()
        self.gate_label.config(text="")
        if m == "window":
            self.win_frame.grid()
        elif m == "prepost":
            self.pp_frame.grid()

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _refresh_ports(self):
        ports = [p.device for p in serial.tools.list_ports.comports()]
        self.port_cb["values"] = ports
        if ports: self.port_cb.set(ports[0])

    def _browse_dir(self):
        d = filedialog.askdirectory(initialdir=self.dir_var.get())
        if d: self.dir_var.set(d)

    def _log(self, msg, tag="info"):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg+"\n", tag)
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0","end")
        self.log_box.configure(state="disabled")
        self._live_index = self._saved_count = 0
        self.stats_label.config(text="Records: 0  |  Saved: 0")

    def _make_path(self, suffix="", ext=".xlsx"):
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        p  = self.prefix_var.get().strip() or "serial_log"
        return os.path.join(self.dir_var.get(), f"{p}{suffix}_{ts}{ext}")

    # ── Connection ────────────────────────────────────────────────────────────

    def _toggle_connection(self):
        if self.logger.serial_port and self.logger.serial_port.is_open:
            self._disconnect()
        else:
            self._connect()

    def _connect(self):
        port = self.port_var.get()
        if not port:
            messagebox.showerror("Error","No port selected."); return
        try:
            baud = int(self.baud_var.get()); hz = float(self.rate_var.get())
            if hz <= 0: raise ValueError
        except ValueError:
            messagebox.showerror("Error","Invalid baud rate or sample rate."); return
        try:
            self.logger.connect(port, baud)
        except Exception as e:
            messagebox.showerror("Connection Error", str(e)); return
        self.logger.start_reading(1.0/hz, self._on_data)
        self.status_label.config(text="● Connected", fg="#4ecca3")
        self.connect_btn.config(text="Disconnect")
        self._log(f"[INFO] Connected to {port} @ {baud} baud | {hz} Hz", "info")
        self._schedule_autosave()

    def _disconnect(self):
        if self.auto_save_job:
            self.after_cancel(self.auto_save_job); self.auto_save_job = None
        if self._cs_running:
            self._stop_continuous_save()
        self.logger.disconnect()
        self.status_label.config(text="● Disconnected", fg="#e94560")
        self.connect_btn.config(text="Connect")
        self._log("[INFO] Disconnected.", "info")

    # ── Data callback ─────────────────────────────────────────────────────────

    def _on_data(self, ts, raw, saved):
        if ts is None:
            self.after(0, self._log, raw, "error"); return
        self.after(0, self._append_live_row, ts, raw, saved)

    def _append_live_row(self, ts, raw, saved):
        self._live_index += 1
        if saved: self._saved_count += 1

        # Determine which colour tag to use
        mode = self.cs_mode_var.get() if self._cs_running else ""
        if saved and mode == "prepost":
            val_tag = "post"          # live post-window
        elif saved:
            val_tag = "win"           # simple window
        else:
            val_tag = "val"

        self.log_box.configure(state="normal")
        self.log_box.insert("end", f"{self._live_index:>5} ", "idx")
        self.log_box.insert("end", f"  {ts.strftime('%H:%M:%S.%f')[:-3]}  ", "ts")
        self.log_box.insert("end", f"{raw}\n", val_tag)
        self.log_box.see("end")
        self.log_box.configure(state="disabled")
        self.stats_label.config(text=f"Records: {self._live_index}  |  Saved: {self._saved_count}")

    # ── Snapshot save ─────────────────────────────────────────────────────────

    def _save_snapshot(self):
        try:
            interval = float(self.interval_var.get())
        except ValueError:
            messagebox.showerror("Error","Invalid snapshot interval."); return
        fp = self._make_path()
        try:
            count = self.logger.save_snapshot_to_excel(fp, interval)
        except Exception as e:
            messagebox.showerror("Save Error", str(e)); return
        if count:
            self._log(f"[SNAPSHOT] {count} records → {fp}", "info")
            messagebox.showinfo("Saved", f"Saved {count} records to:\n{fp}")
        else:
            messagebox.showwarning("No Data","No data in the specified interval.")

    # ── Continuous save ───────────────────────────────────────────────────────

    def _start_continuous_save(self):
        if not (self.logger.serial_port and self.logger.serial_port.is_open):
            messagebox.showwarning("Not Connected","Connect to a serial port first."); return
        if self._cs_running: return

        mode = self.cs_mode_var.get()
        window = period = pre = post = 0.0

        try:
            if mode == "window":
                window = float(self.cs_window_var.get())
                period = float(self.cs_period_var.get())
                if window <= 0 or period <= 0 or window >= period:
                    raise ValueError("window must be > 0 and < period")
            elif mode == "prepost":
                period = float(self.cs_pp_period_var.get())
                pre    = float(self.cs_pre_var.get())
                post   = float(self.cs_post_var.get())
                if period <= 0: raise ValueError("period must be > 0")
                if pre < 0 or post < 0: raise ValueError("pre/post must be >= 0")
                if pre == 0 and post == 0: raise ValueError("set at least one of pre or post > 0")
                if (pre + post) > period:
                    raise ValueError(
                        f"Pre ({pre}s) + Post ({post}s) = {pre+post}s exceeds period ({period}s).\n"
                        "Reduce pre/post or increase period."
                    )
        except ValueError as e:
            messagebox.showerror("Invalid Settings", str(e)); return

        csv_path = self._make_path("_continuous", ext=".csv")
        try:
            self.logger.start_continuous_save(
                csv_path, window=window, period=period, pre=pre, post=post,
                on_gate_cb=self._on_gate_change
            )
        except Exception as e:
            messagebox.showerror("Error", str(e)); return

        self._cs_running = True
        # Apply nth-sample setting
        try:
            nth = int(self.nth_var.get())
            self.logger._nth_step = max(1, nth)
        except ValueError:
            self.logger._nth_step = 1
        self.logger._nth_counter = 0
        self.cs_start_btn.grid_remove()
        self.cs_stop_btn.grid()

        if mode == "all":
            desc = "every reading"
        elif mode == "window":
            desc = f"{window:g}s window / {period:g}s period"
            self._show_gate("open")
        else:
            desc = f"pre={pre:g}s | post={post:g}s | every {period:g}s"
            self._show_gate("wait")

        self.cs_status.config(text=f"● {desc}\n{os.path.basename(csv_path)}")
        self._log(f"[CS START] {desc} → {csv_path}", "info")

    def _stop_continuous_save(self):
        if not self._cs_running: return
        csv_path = self.logger.stop_continuous_save()
        self._cs_running = False
        self.logger._nth_step = 1
        self.logger._nth_counter = 0
        self.cs_stop_btn.grid_remove(); self.cs_start_btn.grid()
        self.cs_status.config(text=""); self.gate_label.config(text="")

        if not csv_path or not os.path.exists(csv_path):
            self._log("[CS STOP] No data recorded.", "info"); return

        xlsx_path = csv_path.replace(".csv", ".xlsx")
        self._log(f"[CS STOP] Converting → {xlsx_path}", "info")

        def _cvt():
            try:
                n = csv_to_excel(csv_path, xlsx_path)
                try: os.remove(csv_path)
                except: pass
                self.after(0, self._cs_done, xlsx_path, n)
            except Exception as e:
                self.after(0, self._log, f"[ERROR] {e}", "error")

        threading.Thread(target=_cvt, daemon=True).start()

    def _cs_done(self, path, count):
        self._log(f"[CS DONE] {count} records → {path}", "info")
        messagebox.showinfo("Done", f"Saved {count} records to:\n{path}")

    def _on_gate_change(self, phase: str):
        self.after(0, self._show_gate, phase)

    def _show_gate(self, phase: str):
        labels = {
            "open":   ("⬤  RECORDING",        "#4ecca3"),
            "closed": ("◯  waiting…",          "#888888"),
            "wait":   ("◯  waiting for trigger","#888888"),
            "pre":    ("⬤  flushing pre-buffer","#4caf50"),
            "post":   ("⬤  POST window",        "#f0a500"),
        }
        text, colour = labels.get(phase, ("", "#888888"))
        self.gate_label.config(text=text, fg=colour)

    # ── Auto-save ─────────────────────────────────────────────────────────────

    def _schedule_autosave(self):
        try:
            every = float(self.autosave_var.get())
        except ValueError:
            every = 0
        if every <= 0: return

        def _do():
            if self.logger.serial_port and self.logger.serial_port.is_open:
                try: iv = float(self.interval_var.get())
                except: iv = 60
                fp = self._make_path("_auto")
                n  = self.logger.save_snapshot_to_excel(fp, iv)
                if n: self._log(f"[AUTO-SAVE] {n} records → {fp}", "info")
                self.auto_save_job = self.after(int(every*1000), _do)
        self.auto_save_job = self.after(int(every*1000), _do)

    def on_close(self):
        self._disconnect(); self.destroy()


# ─────────────────────────────────────────────────────────────────────────────
#  Entry point
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = App()
    app.protocol("WM_DELETE_WINDOW", app.on_close)
    app.mainloop()

